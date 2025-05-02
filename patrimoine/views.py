from django.db.models import Sum, Count, Q, F
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView, TemplateView
from django.urls import reverse_lazy
from django.views.decorators.csrf import csrf_exempt


# api/views.py
from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend

from ..serializers import (BienSerializer, BienDetailSerializer, 
                         CategorieSerializer, EntiteSerializer,
                         StatistiquesSerializer)

# Ajoutez cette ligne pour importer Workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Bien, Categorie, SousCategorie, Entite, HistoriqueValeur, Province, Commune
from .forms import (
    BienForm,
    HistoriqueValeurForm,
    ProfilVehiculeForm,
    ProfilImmeubleForm,
    ProfilInformatiqueForm,
    ProfilEquipementMedicalForm,
    ProfilMobilierForm,
    ProfilTerrainForm,
    ProfilConsommableForm,
)

from collections import defaultdict
from django.db.models.functions import ExtractYear
from django.utils.timezone import now



def accueil_view(request):
    context = {
        "now": now()  # pour afficher l'année dans le footer
    }
    return render(request, "home.html", context)


# --- Mapping centralisé code -> (FormClass, template) ---
# Mapping unifié basé sur vos codes en base
# --- Mapping centralisé code -> (FormClass, template) ---
# Mapping unifié basé sur vos codes en base
FORM_MAPPING = {
    'batiments_administratifs': {'form': ProfilImmeubleForm, 'template': 'patrimoine/_form_immeuble.html'},
    'infrastructures_sanitaires': {'form': ProfilImmeubleForm, 'template': 'patrimoine/_form_immeuble.html'},
    'terrains': {'form': ProfilTerrainForm, 'template': 'patrimoine/_form_terrain.html'},
    'installations_techniques': {'form': ProfilImmeubleForm, 'template': 'patrimoine/_form_immeuble.html'},
    'equipement_medical': {'form': ProfilEquipementMedicalForm, 'template': 'patrimoine/_form_equipement_medical.html'},
    'materiel_informatique': {'form': ProfilInformatiqueForm, 'template': 'patrimoine/_form_informatique.html'},
    'mobilier_de_bureau': {'form': ProfilMobilierForm, 'template': 'patrimoine/_form_mobilier.html'},
    'vehicules_de_service': {'form': ProfilVehiculeForm, 'template': 'patrimoine/_form_vehicule.html'},
    'consommables_de_valeur_et_stocks_strategiques': {
        'form': ProfilConsommableForm,
        'template': 'patrimoine/_form_consommables.html',
    },
}


# --- Class-based views ---
class BienListView(ListView):
    model = Bien
    template_name = 'patrimoine/bien_list.html'
    context_object_name = 'biens'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().select_related('categorie', 'entite').order_by('nom')
        search_query = self.request.GET.get('q')
        if search_query:
            queryset = queryset.filter(
                Q(nom__icontains=search_query) | Q(categorie__nom__icontains=search_query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('q', '')
        return context


class BienCreateView(CreateView):
    model = Bien
    form_class = BienForm
    template_name = 'patrimoine/bien_form.html'
    success_url = reverse_lazy('biens:bien_list')


class BienUpdateView(UpdateView):
    model = Bien
    form_class = BienForm
    template_name = 'patrimoine/bien_form.html'
    success_url = reverse_lazy('biens:bien_list')


class BienDeleteView(DeleteView):
    model = Bien
    template_name = 'patrimoine/bien_confirm_delete.html'
    success_url = reverse_lazy('biens:bien_list')


class DetailViewMixin:
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['historiques'] = self.object.historiques.order_by('date')
        return context

# views.py - Exemple de vue optimisée
from django.views.generic import DetailView
from .models import Bien
from .services.bien_service import BienService
from .permissions.patrimoine_permissions import bien_permission_required, BienPermissions

class BienDetailView(DetailView):
    model = Bien
    template_name = 'patrimoine/bien_detail.html'
    context_object_name = 'bien'
    
    def dispatch(self, request, *args, **kwargs):
        # Vérifier les permissions
        bien = self.get_object()
        if not BienPermissions.peut_voir_bien(request.user, bien):
            raise PermissionDenied("Vous n'avez pas les permissions pour voir ce bien.")
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        # Optimiser les requêtes avec select_related et prefetch_related
        return Bien.objects.select_related(
            'categorie', 
            'sous_categorie',
            'entite',
            'commune__departement__province'
        ).prefetch_related(
            'historiques',
            'responsabilites__responsable'
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Utiliser le service pour obtenir des données
        context['historiques'] = self.object.historiques.order_by('-date')
        context['form'] = HistoriqueValeurForm()
        context['all_responsables'] = ResponsableBien.objects.all().order_by('nom', 'prenom')
        
        # Charger les profils techniques
        self._charger_profil_technique(context)
        
        return context
    
    def _charger_profil_technique(self, context):
        """Extrait la logique de chargement des profils techniques"""
        bien = self.object
        profil_attrs = [
            'profil_vehicule', 'profil_immeuble', 'profil_informatique', 
            'profil_equipement_medical', 'profil_mobilier', 'profil_terrain', 
            'profil_consommable'
        ]
        
        for attr in profil_attrs:
            if hasattr(bien, attr):
                context[attr] = getattr(bien, attr)
                
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        action = request.POST.get('action') or request.GET.get('action')
        
        # Traiter l'assignation d'un responsable
        if action == 'assigner_responsable':
            return self.assigner_responsable(request)
        
        # Par défaut, traiter l'ajout d'une valeur historique
        return self.ajouter_valeur_historique(request)
    
    def ajouter_valeur_historique(self, request):
        form = HistoriqueValeurForm(request.POST)
        if form.is_valid():
            historique = form.save(commit=False)
            historique.bien = self.object
            historique.save()
            messages.success(request, "La valeur a été ajoutée avec succès.")
            return redirect('biens:bien_detail', pk=self.object.pk)
        
        # En cas d'erreur, recharger la page avec le formulaire contenant les erreurs
        context = self.get_context_data()
        context['form'] = form
        messages.error(request, "Erreur lors de l'ajout de la valeur. Veuillez vérifier les champs.")
        return render(request, self.template_name, context)
    
    def assigner_responsable(self, request):
        responsable_id = request.POST.get('responsable')
        type_affectation = request.POST.get('type_affectation')
        date_affectation = request.POST.get('date_affectation')
        motif = request.POST.get('motif', '')
        
        # Valider les champs requis
        if not (responsable_id and type_affectation and date_affectation):
            messages.error(request, "Veuillez remplir tous les champs obligatoires.")
            return redirect('biens:bien_detail', pk=self.object.pk)
        
        try:
            responsable = ResponsableBien.objects.get(pk=responsable_id)
            
            # Créer la responsabilité
            responsabilite = BienResponsabilite(
                bien=self.object,
                responsable=responsable,
                date_affectation=date_affectation,
                type_affectation=type_affectation,
                motif=motif
            )
            responsabilite.save()
            
            messages.success(request, f"Le bien a été assigné à {responsable.prenom} {responsable.nom} avec succès.")
        except ResponsableBien.DoesNotExist:
            messages.error(request, "Le responsable sélectionné n'existe pas.")
        except Exception as e:
            messages.error(request, f"Une erreur est survenue lors de l'assignation: {str(e)}")
        
        return redirect('biens:bien_detail', pk=self.object.pk)

class DashboardView(TemplateView):
    template_name = 'patrimoine/dashboard.html'

    def get(self, request, *args, **kwargs):
        # Si format=excel est dans les paramètres, générer un fichier Excel
        if request.GET.get('format') == 'excel':
            return self.export_excel(request)
        
        # Sinon, continuer avec le comportement normal
        return super().get(request, *args, **kwargs)
    
    def export_excel(self, request):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Tableau de bord"
        worksheet.append(['Catégorie', 'Nombre de biens', 'Valeur totale (FCFA)'])

        # Appliquer les filtres
        all_biens = self.get_filtered_biens(request)

        biens_par_categorie = all_biens.values('categorie__nom').annotate(
            nb_biens=Count('id'), total=Sum('valeur_initiale')
        )

        for category_data in biens_par_categorie:
            worksheet.append([
                category_data['categorie__nom'],
                category_data['nb_biens'],
                float(category_data['total'] or 0),
            ])

        for col_idx in range(1, 4):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 25

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=dashboard.xlsx'
        workbook.save(response)
        return response
    
    def get_filtered_biens(self, request):
        all_biens = Bien.objects.select_related('commune', 'categorie', 'entite')
        selected_year = request.GET.get('annee')
        selected_commune = request.GET.get('commune')

        if selected_year:
            all_biens = all_biens.filter(date_acquisition__year=selected_year)
        if selected_commune:
            all_biens = all_biens.filter(commune__id=selected_commune)
            
        return all_biens

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Utilisez la méthode get_filtered_biens pour appliquer les filtres
        all_biens = self.get_filtered_biens(self.request)

        context.update({
            'valeur_totale': all_biens.aggregate(Sum('valeur_initiale'))['valeur_initiale__sum'] or 0,
            'par_categorie': all_biens.values('categorie__nom').annotate(
                nb_biens=Count('id'), total=Sum('valeur_initiale')
            ),
            'par_entite': all_biens.values('entite__nom').annotate(
                nb_biens=Count('id'), total=Sum('valeur_initiale')
            ),
            'par_commune': all_biens.values('commune__nom').annotate(
                nb_biens=Count('id'), total=Sum('valeur_initiale')
            ),
            'annees_disponibles': Bien.objects.annotate(annee=ExtractYear('date_acquisition'))
                .values_list('annee', flat=True)
                .distinct()
                .order_by('-annee'),
            'communes': Commune.objects.all(),
        })

        return context
class CarteView(TemplateView):
    template_name = 'patrimoine/carte.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        biens = Bien.objects.select_related(
            'commune__departement__province',
            'categorie',
            'entite'
        )

        carte_data = []

        for bien in biens:
            commune = bien.commune
            if commune and commune.latitude and commune.longitude:
                carte_data.append({
                    'commune_nom': commune.nom,
                    'latitude': float(commune.latitude),
                    'longitude': float(commune.longitude),
                    'categorie': bien.categorie.nom if bien.categorie else '',
                    'annee': bien.date_acquisition.year if bien.date_acquisition else '',
                    'province_nom': commune.departement.province.nom if commune.departement and commune.departement.province else '',
                    'entite_nom': bien.entite.nom if bien.entite else '',
                    'nb_biens': 1,
                    'total': float(bien.valeur_initiale) if bien.valeur_initiale else 0,
                })

        context['carte_data'] = carte_data
        context['categories'] = sorted({bien.categorie.nom for bien in biens if bien.categorie})
        context['annees'] = sorted({bien.date_acquisition.year for bien in biens if bien.date_acquisition}, reverse=True)
        context['provinces'] = list(Province.objects.values_list('nom', flat=True))
        context['entites'] = list(Entite.objects.values_list('nom', flat=True))

        return context
@csrf_exempt
def get_profil_form(request):
    sous_categorie_id = request.GET.get('sous_categorie_id')
    if not sous_categorie_id or not sous_categorie_id.isdigit():
        return JsonResponse({'error': 'ID invalide'}, status=400)
    try:
        sous_categorie = SousCategorie.objects.get(pk=int(sous_categorie_id))
        form_config = FORM_MAPPING.get(sous_categorie.code)
        if form_config and form_config['form']:
            html_form = render_to_string(
                form_config['template'], {'profil_form': form_config['form']()}, request=request
            )
            return JsonResponse({'form': html_form})
        return JsonResponse({'form': ''})
    except SousCategorie.DoesNotExist:
        return JsonResponse({'error': 'Sous-catégorie introuvable'}, status=404)

def load_sous_categories(request):
    categorie_id = request.GET.get('categorie')
    sous_categories = SousCategorie.objects.filter(categorie_id=categorie_id).order_by('nom')
    html_dropdown = render_to_string('patrimoine/sous_categorie_dropdown_list.html', {'sous_categories': sous_categories})
    return HttpResponse(html_dropdown)

@csrf_exempt
def ajouter_bien(request):
    bien_form = BienForm(request.POST or None, request.FILES or None)
    profil_form = None

    if request.method == 'POST' and bien_form.is_valid():
        bien = bien_form.save(commit=False)
        sous_categorie_code = bien_form.cleaned_data['sous_categorie'].code
        profil_form_class = FORM_MAPPING.get(sous_categorie_code, {}).get('form')
        profil_valid = True

        if profil_form_class:
            profil_form = profil_form_class(request.POST, request.FILES)
            profil_valid = profil_form.is_valid()

        bien.save()

        if profil_form_class and profil_valid:
            profil = profil_form.save(commit=False)
            profil.bien = bien
            profil.save()

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        return redirect('biens:bien_detail', pk=bien.pk)

    return render(request, 'patrimoine/ajouter_bien.html', {'bien_form': bien_form, 'profil_form': profil_form})


# api/views.py
class BienViewSet(viewsets.ModelViewSet):
    queryset = Bien.objects.all()
    serializer_class = BienSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['categorie', 'sous_categorie', 'entite', 'commune__departement__province']
    search_fields = ['nom', 'description', 'numero_serie']
    ordering_fields = ['nom', 'valeur_initiale', 'date_acquisition']
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return BienDetailSerializer
        return BienSerializer
    
    @action(detail=False, methods=['get'])
    def statistiques(self, request):
        """Fournit des statistiques globales sur les biens"""
        # Filtrer selon les paramètres
        queryset = self.filter_queryset(self.get_queryset())
        
        # Calculer les statistiques
        stats = {
            'nombre_total': queryset.count(),
            'valeur_totale': queryset.aggregate(Sum('valeur_initiale'))['valeur_initiale__sum'] or 0,
            'par_categorie': list(queryset.values(
                categorie=F('categorie__nom')
            ).annotate(
                count=Count('id'),
                valeur=Sum('valeur_initiale')
            ).order_by('-count')),
            'par_entite': list(queryset.values(
                entite=F('entite__nom')
            ).annotate(
                count=Count('id'),
                valeur=Sum('valeur_initiale')
            ).order_by('-count')),
        }
        
        serializer = StatistiquesSerializer(stats)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def historique(self, request, pk=None):
        """Récupère l'historique des valeurs d'un bien"""
        bien = self.get_object()
        historiques = bien.historiques.all().order_by('-date')
        
        from ..serializers import HistoriqueValeurSerializer
        serializer = HistoriqueValeurSerializer(historiques, many=True)
        return Response(serializer.data)